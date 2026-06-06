import math
import os
import random
import re
import zipfile
from collections import deque
from datetime import datetime
from io import BytesIO
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows


NEW_TEACHER_YEARS = 2
INVIGILATOR_ATTEMPTS = 40

GRADUATE_ID_COL = "身份证号码"
GRADUATE_MAJOR_COL = "专业"
GRADUATE_LEVEL_COL = "培养层次"
SCORE_ID_COL = "ks_sfz"
SCORE_VALUE_COL = "zf"
PASS_SCORE = 425
TARGET_LEVELS = ["本科", "专科"]
EXCLUDE_MAJORS = [
    "舞蹈学(本)", "音乐学(本)", "音乐表演(本)",
    "社会体育指导与管理(本)", "体育教育(本)",
    "公共艺术(本)", "环境设计(本)", "美术学(本)", "英语(本)",
]


def _require_columns(df: pd.DataFrame, required: set[str], label: str):
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{label} 缺少必要列：{', '.join(sorted(missing))}")


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip()
    return (cleaned or "Sheet")[:31]


def get_classroom_name(room_name: str) -> str:
    if "-" in str(room_name):
        return str(room_name).split("-")[0]
    return str(room_name)


class _MinCostMaxFlow:
    def __init__(self, size: int):
        self.graph = [[] for _ in range(size)]

    def add_edge(self, start: int, end: int, capacity: int, cost: int):
        forward = [end, capacity, cost, None]
        backward = [start, 0, -cost, forward]
        forward[3] = backward
        self.graph[start].append(forward)
        self.graph[end].append(backward)

    def min_cost_flow(self, source: int, sink: int, required_flow: int) -> tuple[int, int]:
        flow = 0
        cost = 0
        node_count = len(self.graph)

        while flow < required_flow:
            dist = [float("inf")] * node_count
            in_queue = [False] * node_count
            previous_node = [-1] * node_count
            previous_edge = [None] * node_count
            dist[source] = 0
            queue = deque([source])
            in_queue[source] = True

            while queue:
                current = queue.popleft()
                in_queue[current] = False
                for edge in self.graph[current]:
                    end, capacity, edge_cost, _ = edge
                    if capacity <= 0 or dist[end] <= dist[current] + edge_cost:
                        continue
                    dist[end] = dist[current] + edge_cost
                    previous_node[end] = current
                    previous_edge[end] = edge
                    if not in_queue[end]:
                        queue.append(end)
                        in_queue[end] = True

            if previous_node[sink] == -1:
                break

            pushed = required_flow - flow
            node = sink
            while node != source:
                edge = previous_edge[node]
                pushed = min(pushed, edge[1])
                node = previous_node[node]

            node = sink
            while node != source:
                edge = previous_edge[node]
                edge[1] -= pushed
                edge[3][1] += pushed
                cost += pushed * edge[2]
                node = previous_node[node]
            flow += pushed

        return flow, cost


def _teacher_year(teacher_id: int) -> int | None:
    match = re.match(r"^(\d{4})", str(teacher_id))
    if not match:
        return None
    return int(match.group(1))


def _is_new_teacher(teacher: dict[str, Any], current_year: int) -> bool:
    year = teacher.get("year")
    return year is not None and year >= current_year - NEW_TEACHER_YEARS + 1


def _experience_gap(first: dict[str, Any], second: dict[str, Any]) -> int:
    first_year = first.get("year")
    second_year = second.get("year")
    if first_year is None or second_year is None:
        return 0
    return abs(first_year - second_year)


def _invigilator_pair_score(jia: dict[str, Any], yi: dict[str, Any], room_index: int, room_count: int, current_year: int) -> int:
    double_new = _is_new_teacher(jia, current_year) and _is_new_teacher(yi, current_year)
    same_college = jia["college"] == yi["college"]
    early_weight = room_count - room_index

    score = _experience_gap(jia, yi) * 1000
    if double_new:
        score -= 100000 + early_weight * 100
    if same_college:
        score -= 5000 + early_weight * 10
    return score


def _format_teacher(teacher: dict[str, Any]) -> str:
    return f"{teacher['name']}({teacher['id']},{teacher['gender']})"


def _assign_yi_by_flow(
    rooms: list[dict[str, Any]],
    jia_assignments: dict[Any, dict[str, Any]],
    yi_candidates: list[dict[str, Any]],
    current_year: int,
) -> tuple[dict[Any, dict[str, Any]] | None, int]:
    room_count = len(rooms)
    source = 0
    female_offset = 1
    room_offset = female_offset + len(yi_candidates)
    sink = room_offset + room_count
    flow = _MinCostMaxFlow(sink + 1)

    for idx, _ in enumerate(yi_candidates):
        flow.add_edge(source, female_offset + idx, 1, 0)

    edge_lookup: dict[tuple[int, int], tuple[Any, dict[str, Any]]] = {}
    for female_idx, yi in enumerate(yi_candidates):
        for room_idx, room in enumerate(rooms):
            room_no = room["room_no"]
            jia = jia_assignments[room_no]
            score = _invigilator_pair_score(jia, yi, room_idx, room_count, current_year)
            start = female_offset + female_idx
            end = room_offset + room_idx
            flow.add_edge(start, end, 1, -score)
            edge_lookup[(start, end)] = (room_no, yi)

    for room_idx, _ in enumerate(rooms):
        flow.add_edge(room_offset + room_idx, sink, 1, 0)

    assigned_flow, cost = flow.min_cost_flow(source, sink, room_count)
    if assigned_flow != room_count:
        return None, cost

    assignments: dict[Any, dict[str, Any]] = {}
    for start in range(female_offset, female_offset + len(yi_candidates)):
        for edge in flow.graph[start]:
            end = edge[0]
            if end < room_offset or end >= sink:
                continue
            # Original forward edges have capacity 0 after carrying one unit of flow.
            if edge[1] == 0 and (start, end) in edge_lookup:
                room_no, yi = edge_lookup[(start, end)]
                assignments[room_no] = yi
                break

    if len(assignments) != room_count:
        return None, cost
    return assignments, cost


def create_invigilator_plan(teachers_bytes: bytes, rooms_bytes: bytes, room_count: int) -> dict[str, Any]:
    teachers_df = pd.read_excel(BytesIO(teachers_bytes))
    rooms_df = pd.read_excel(BytesIO(rooms_bytes))
    _require_columns(teachers_df, {"id", "name", "gender", "college"}, "监考员表")
    _require_columns(rooms_df, {"room_no", "room_name"}, "考场表")

    if room_count < 1:
        raise ValueError("考场数量必须大于 0")
    if room_count > len(rooms_df):
        raise ValueError(f"考场数量不能超过考场表总数 {len(rooms_df)}")

    teachers = []
    for _, row in teachers_df.iterrows():
        teacher_id = int(row["id"])
        teachers.append({
            "id": teacher_id,
            "name": str(row["name"]).strip(),
            "gender": str(row["gender"]).strip(),
            "college": str(row["college"]).strip(),
            "year": _teacher_year(teacher_id),
        })

    rooms = rooms_df.head(room_count).to_dict("records")
    for room in rooms:
        room["classroom"] = get_classroom_name(room["room_name"])

    female_teachers = [t for t in teachers if t["gender"] == "女"]
    male_teachers = [t for t in teachers if t["gender"] == "男"]
    classrooms: dict[str, list[dict[str, Any]]] = {}
    for room in rooms:
        classrooms.setdefault(room["classroom"], []).append(room)

    if len(female_teachers) < len(rooms):
        raise ValueError("女性监考人数不足，无法满足所有考场乙位要求")
    if len(teachers) < len(rooms) * 2:
        raise ValueError("监考员总人数不足，无法满足每个考场两名监考员要求")
    if len(male_teachers) < len(classrooms):
        raise ValueError(f"男性监考人数不足，无法满足每个实体教室至少一名男性要求（需 {len(classrooms)} 人，现有 {len(male_teachers)} 人）")

    current_year = datetime.now().year
    jia_count = len(rooms)
    male_jia_count = min(len(male_teachers), jia_count)
    female_jia_count = jia_count - male_jia_count
    if len(female_teachers) - female_jia_count < len(rooms):
        raise ValueError("女性监考人数不足，无法同时满足乙位女性和甲位人员数量要求")

    coverage_slots = [classroom_rooms[0]["room_no"] for classroom_rooms in classrooms.values()]
    coverage_slot_set = set(coverage_slots)
    flexible_slots = [room["room_no"] for room in rooms if room["room_no"] not in coverage_slot_set]
    if female_jia_count > len(flexible_slots):
        raise ValueError("男性监考人数不足，无法在所有实体教室完成男性覆盖")

    best_result: dict[str, Any] | None = None
    for attempt in range(INVIGILATOR_ATTEMPTS):
        male_pool = male_teachers[:]
        female_pool = female_teachers[:]
        random.shuffle(male_pool)
        random.shuffle(female_pool)

        selected_female_jia = female_pool[:female_jia_count]
        yi_candidates = female_pool[female_jia_count:]
        selected_male_jia = male_pool[:male_jia_count]
        if len(yi_candidates) < len(rooms):
            continue

        jia_assignments: dict[Any, dict[str, Any]] = {}
        random.shuffle(coverage_slots)
        random.shuffle(flexible_slots)

        male_cursor = 0
        for room_no in coverage_slots:
            jia_assignments[room_no] = selected_male_jia[male_cursor]
            male_cursor += 1

        remaining_jia = selected_male_jia[male_cursor:] + selected_female_jia
        random.shuffle(remaining_jia)
        for room_no, teacher in zip(flexible_slots, remaining_jia):
            jia_assignments[room_no] = teacher

        if len(jia_assignments) != len(rooms):
            continue

        yi_assignments, flow_cost = _assign_yi_by_flow(rooms, jia_assignments, yi_candidates, current_year)
        if yi_assignments is None:
            continue

        violations = []
        score = -flow_cost
        for idx, room in enumerate(rooms):
            room_no = room["room_no"]
            jia = jia_assignments[room_no]
            yi = yi_assignments[room_no]
            if _is_new_teacher(jia, current_year) and _is_new_teacher(yi, current_year):
                violations.append({
                    "考场号": room_no,
                    "考场名称": room["room_name"],
                    "违规/降级类型": "双新晋搭配",
                    "说明": "甲乙均为两年内新晋教职工",
                    "监考员甲": _format_teacher(jia),
                    "监考员乙": _format_teacher(yi),
                })
            if jia["college"] == yi["college"]:
                violations.append({
                    "考场号": room_no,
                    "考场名称": room["room_name"],
                    "违规/降级类型": "甲乙同学院",
                    "说明": "为满足更高优先级约束，放宽学院均衡偏好",
                    "监考员甲": _format_teacher(jia),
                    "监考员乙": _format_teacher(yi),
                })

        candidate = {
            "score": score,
            "jia_assignments": jia_assignments,
            "yi_assignments": yi_assignments,
            "violations": violations,
        }
        if best_result is None:
            best_result = candidate
            continue
        current_double_new = sum(1 for item in violations if item["违规/降级类型"] == "双新晋搭配")
        best_double_new = sum(1 for item in best_result["violations"] if item["违规/降级类型"] == "双新晋搭配")
        current_same_college = sum(1 for item in violations if item["违规/降级类型"] == "甲乙同学院")
        best_same_college = sum(1 for item in best_result["violations"] if item["违规/降级类型"] == "甲乙同学院")
        if (
            current_double_new,
            current_same_college,
            -score,
        ) < (
            best_double_new,
            best_same_college,
            -best_result["score"],
        ):
            best_result = candidate

    if best_result is None:
        raise ValueError("无法在当前硬约束下生成完整分配，请检查女性人数、男性人数和考场数量")

    rows = []
    male_covered_classrooms = set()
    year_gaps = []
    same_college_count = 0
    double_new_count = 0
    for room in rooms:
        room_no = room["room_no"]
        jia = best_result["jia_assignments"][room_no]
        yi = best_result["yi_assignments"][room_no]
        if jia["gender"] == "男":
            male_covered_classrooms.add(room["classroom"])
        year_gap = _experience_gap(jia, yi)
        year_gaps.append(year_gap)
        same_college = jia["college"] == yi["college"]
        double_new = _is_new_teacher(jia, current_year) and _is_new_teacher(yi, current_year)
        same_college_count += int(same_college)
        double_new_count += int(double_new)
        rows.append({
            "考场号": room_no,
            "考场名称": room["room_name"],
            "实体教室": room["classroom"],
            "jia": {
                "id": jia["id"],
                "name": jia["name"],
                "gender": jia["gender"],
                "college": jia["college"],
                "year": jia.get("year"),
            },
            "yi": {
                "id": yi["id"],
                "name": yi["name"],
                "gender": yi["gender"],
                "college": yi["college"],
                "year": yi.get("year"),
            },
            "工号年份差": year_gap,
            "是否双新晋": "是" if double_new else "否",
            "是否同学院": "是" if same_college else "否",
        })

    missing_male_coverage = sorted(set(classrooms) - male_covered_classrooms)
    violations = best_result["violations"][:]
    for classroom in missing_male_coverage:
        violations.append({
            "考场号": "",
            "考场名称": classroom,
            "违规/降级类型": "实体教室缺少男性",
            "说明": "硬约束未满足，请人工调整",
            "监考员甲": "",
            "监考员乙": "",
        })

    summary = {
        "room_count": len(rooms),
        "classroom_count": len(classrooms),
        "teacher_count": len(teachers),
        "female_count": len(female_teachers),
        "male_count": len(male_teachers),
        "double_new_count": double_new_count,
        "same_college_count": same_college_count,
        "avg_year_gap": round(sum(year_gaps) / len(year_gaps), 2) if year_gaps else 0,
        "new_teacher_threshold": current_year - NEW_TEACHER_YEARS + 1,
        "male_coverage_ok": not missing_male_coverage,
    }
    return {
        "rows": rows,
        "summary": summary,
        "violations": violations,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


def _invigilator_plan_to_detailed_workbook(plan: dict[str, Any]) -> BytesIO:
    rows = plan.get("rows") or []
    summary = plan.get("summary") or {}
    arrangements = []
    for row in rows:
        jia = row["jia"]
        yi = row["yi"]
        arrangements.append({
            "考场号": row["考场号"],
            "考场名称": row["考场名称"],
            "实体教室": row["实体教室"],
            "监考员甲": _format_teacher(jia),
            "甲学院": jia["college"],
            "甲入职年份": jia.get("year") or "",
            "监考员乙": _format_teacher(yi),
            "乙学院": yi["college"],
            "乙入职年份": yi.get("year") or "",
            "工号年份差": row.get("工号年份差", ""),
            "是否双新晋": row.get("是否双新晋", ""),
            "是否同学院": row.get("是否同学院", ""),
        })

    summary_rows = [
        {"项目": "考场数", "值": summary.get("room_count", 0)},
        {"项目": "实体教室数", "值": summary.get("classroom_count", 0)},
        {"项目": "监考员总数", "值": summary.get("teacher_count", 0)},
        {"项目": "女性监考员数", "值": summary.get("female_count", 0)},
        {"项目": "男性监考员数", "值": summary.get("male_count", 0)},
        {"项目": "乙位女性硬约束", "值": "已满足"},
        {"项目": "每名老师只出现一次", "值": "已满足"},
        {"项目": "实体教室男性覆盖", "值": "已满足" if summary.get("male_coverage_ok") else "未满足"},
        {"项目": "双新晋搭配数", "值": summary.get("double_new_count", 0)},
        {"项目": "甲乙同学院数", "值": summary.get("same_college_count", 0)},
        {"项目": "平均工号年份差", "值": summary.get("avg_year_gap", 0)},
        {"项目": "两年内新晋判定", "值": f"{summary.get('new_teacher_threshold', '')} 年及以后"},
    ]

    output = BytesIO()
    output.invigilator_summary = summary
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(arrangements).to_excel(writer, index=False, sheet_name="分配结果")
        violations_df = pd.DataFrame(plan.get("violations") or [{
            "考场号": "",
            "考场名称": "",
            "违规/降级类型": "无",
            "说明": "所有软约束均已满足",
            "监考员甲": "",
            "监考员乙": "",
        }])
        violations_df.to_excel(writer, index=False, sheet_name="违规降级说明")
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="统计摘要")
    output.seek(0)
    return output


def assign_invigilators(teachers_bytes: bytes, rooms_bytes: bytes, room_count: int) -> BytesIO:
    return _invigilator_plan_to_detailed_workbook(
        create_invigilator_plan(teachers_bytes, rooms_bytes, room_count)
    )


def _teacher_export_name(teacher: dict[str, Any], duplicate_names: set[str]) -> str:
    name = str(teacher.get("name") or "").strip()
    if name in duplicate_names:
        return f"{name}({teacher.get('id')})"
    return name


def export_invigilator_final(rows: list[dict[str, Any]]) -> BytesIO:
    name_counts: dict[str, int] = {}
    for row in rows:
        for role in ("jia", "yi"):
            name = str((row.get(role) or {}).get("name") or "").strip()
            if name:
                name_counts[name] = name_counts.get(name, 0) + 1
    duplicate_names = {name for name, count in name_counts.items() if count > 1}

    export_rows = []
    for row in rows:
        export_rows.append({
            "考场号": row.get("考场号"),
            "考场名称": row.get("考场名称"),
            "监考员甲": _teacher_export_name(row.get("jia") or {}, duplicate_names),
            "监考员乙": _teacher_export_name(row.get("yi") or {}, duplicate_names),
        })

    output = BytesIO()
    pd.DataFrame(export_rows).to_excel(output, index=False)
    output.seek(0)
    return output


def generate_seat_labels_pdf(num_rooms: int, num_seats: int = 30, cols: int = 3, rows: int = 10, font_size: int = 40) -> BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("缺少 reportlab 依赖，请先安装 requirements.txt") from exc

    if num_rooms < 1 or num_seats < 1:
        raise ValueError("考场数量和座位数必须大于 0")
    if cols < 1 or rows < 1:
        raise ValueError("列数和行数必须大于 0")

    page_width, page_height = A4
    left_margin = right_margin = top_margin = bottom_margin = 10 * mm
    h_spacing = v_spacing = 2 * mm
    usable_width = page_width - left_margin - right_margin - h_spacing * (cols - 1)
    usable_height = page_height - top_margin - bottom_margin - v_spacing * (rows - 1)
    label_width = usable_width / cols
    label_height = usable_height / rows
    rooms_per_page = cols * rows

    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)

    for seat in range(1, num_seats + 1):
        pages_needed = math.ceil(num_rooms / rooms_per_page)
        for page in range(pages_needed):
            start_room = page * rooms_per_page + 1
            rooms_this_page = min(num_rooms - page * rooms_per_page, rooms_per_page)
            for col in range(cols):
                for row in range(rows):
                    idx = row + col * rows
                    if idx >= rooms_this_page:
                        continue
                    room = start_room + idx
                    x = left_margin + col * (label_width + h_spacing)
                    y = page_height - top_margin - (row + 1) * label_height - row * v_spacing
                    pdf.setLineWidth(0.8)
                    pdf.setStrokeColor(colors.black)
                    pdf.rect(x, y, label_width, label_height)
                    pdf.setFont("Helvetica", font_size)
                    pdf.drawCentredString(
                        x + label_width / 2,
                        y + label_height / 2 - font_size / 2.8,
                        f"{room}-{seat:02d}",
                    )
            pdf.showPage()

    pdf.save()
    output.seek(0)
    return output


SEAT_LABEL_REQUIRED_COLUMNS = ["考场号", "座位号", "姓名"]


def _read_excel_text(file_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(BytesIO(file_bytes), dtype=str).fillna("")


def get_seat_label_columns(file_bytes: bytes) -> list[str]:
    df = _read_excel_text(file_bytes)
    return [str(col).strip() for col in df.columns if str(col).strip()]


def _clean_cell(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _parse_positive_int(value: Any) -> int | None:
    text = _clean_cell(value)
    if re.fullmatch(r"\d+", text):
        parsed = int(text)
        return parsed if parsed > 0 else None
    return None


def _seat_key(room_no: int, seat_no: int) -> tuple[int, int]:
    return room_no, seat_no


def validate_seat_label_roster(
    file_bytes: bytes,
    id_column: str,
    standard_seats: int = 30,
) -> dict:
    df = _read_excel_text(file_bytes)
    df.columns = [str(col).strip() for col in df.columns]
    id_column = (id_column or "").strip()
    errors = []
    warnings = []
    stats = []
    records = {}

    missing = [col for col in SEAT_LABEL_REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        errors.append({
            "type": "缺少必要列",
            "room_no": "",
            "seat_no": "",
            "names": "",
            "message": f"缺少必要列：{'、'.join(missing)}",
        })
    if not id_column:
        errors.append({
            "type": "未选择编号字段",
            "room_no": "",
            "seat_no": "",
            "names": "",
            "message": "请选择一个编号字段",
        })
    elif id_column not in df.columns:
        errors.append({
            "type": "编号字段不存在",
            "room_no": "",
            "seat_no": "",
            "names": "",
            "message": f"编号字段不存在：{id_column}",
        })

    if errors:
        return _build_seat_precheck_result(errors, warnings, stats, records, standard_seats)

    raw_records_by_key = {}
    room_seats = {}

    for idx, row in df.iterrows():
        excel_row = idx + 2
        room_raw = _clean_cell(row.get("考场号"))
        seat_raw = _clean_cell(row.get("座位号"))
        name = _clean_cell(row.get("姓名"))
        identifier = _clean_cell(row.get(id_column))
        room_no = _parse_positive_int(room_raw)
        seat_no = _parse_positive_int(seat_raw)

        if room_no is None:
            errors.append({
                "type": "考场号不合法",
                "room_no": room_raw,
                "seat_no": seat_raw,
                "names": name,
                "message": f"第 {excel_row} 行考场号不是正整数：{room_raw or '空'}",
            })
            continue
        if seat_no is None or seat_no > standard_seats:
            errors.append({
                "type": "座位号不合法",
                "room_no": room_no,
                "seat_no": seat_raw,
                "names": name,
                "message": f"第 {excel_row} 行座位号必须为 1-{standard_seats}：{seat_raw or '空'}",
            })
            continue
        if not name:
            errors.append({
                "type": "姓名为空",
                "room_no": room_no,
                "seat_no": seat_no,
                "names": "",
                "message": f"第 {excel_row} 行姓名为空",
            })
        if not identifier:
            errors.append({
                "type": "编号为空",
                "room_no": room_no,
                "seat_no": seat_no,
                "names": name,
                "message": f"第 {excel_row} 行 {id_column} 为空",
            })

        key = _seat_key(room_no, seat_no)
        raw_records_by_key.setdefault(key, []).append({
            "excel_row": excel_row,
            "room_no": room_no,
            "seat_no": seat_no,
            "name": name,
            "identifier": identifier,
        })
        room_seats.setdefault(room_no, set()).add(seat_no)

    for (room_no, seat_no), rows in raw_records_by_key.items():
        if len(rows) > 1:
            errors.append({
                "type": "重复座位",
                "room_no": room_no,
                "seat_no": f"{seat_no:02d}",
                "names": "、".join(r["name"] or f"第{r['excel_row']}行" for r in rows),
                "message": f"{room_no}考场{seat_no:02d}号出现 {len(rows)} 条记录",
            })
        elif rows[0]["name"] and rows[0]["identifier"]:
            records[(room_no, seat_no)] = rows[0]

    if room_seats:
        room_numbers = sorted(room_seats)
        expected_rooms = set(range(1, room_numbers[-1] + 1))
        missing_rooms = sorted(expected_rooms - set(room_numbers))
        if room_numbers[0] != 1 or missing_rooms:
            if room_numbers[0] != 1:
                missing_rooms = sorted(set(range(1, room_numbers[0])) | set(missing_rooms))
            errors.append({
                "type": "考场号不连续",
                "room_no": "",
                "seat_no": "",
                "names": "",
                "message": f"考场号必须从 1 连续到最大考场号，缺少：{', '.join(map(str, missing_rooms))}",
            })

        for room_no in room_numbers:
            seats = sorted(room_seats[room_no])
            if not seats:
                continue
            expected_seats = set(range(1, seats[-1] + 1))
            missing_seats = sorted(expected_seats - set(seats))
            if seats[0] != 1 or missing_seats:
                if seats[0] != 1:
                    missing_seats = sorted(set(range(1, seats[0])) | set(missing_seats))
                errors.append({
                    "type": "座位号不连续",
                    "room_no": room_no,
                    "seat_no": "",
                    "names": "",
                    "message": f"{room_no}考场座位号中间断号，缺少：{', '.join(f'{s:02d}' for s in missing_seats)}",
                })

        for room_no in room_numbers:
            seats = sorted(room_seats[room_no])
            count = len(seats)
            max_seat = seats[-1] if seats else 0
            stats.append({
                "room_no": room_no,
                "count": count,
                "seat_range": f"1-{max_seat}" if max_seat else "",
                "full": count == standard_seats,
                "remark": "" if count == standard_seats else f"不满员，{count}人",
            })

        seen_underfilled = None
        for stat in stats:
            if stat["count"] < standard_seats and seen_underfilled is None:
                seen_underfilled = stat
            elif stat["count"] == standard_seats and seen_underfilled is not None:
                warnings.append({
                    "type": "不满员考场后存在满员考场",
                    "room_no": seen_underfilled["room_no"],
                    "seat_no": "",
                    "names": "",
                    "message": f"{seen_underfilled['room_no']}考场只有{seen_underfilled['count']}人，但后续 {stat['room_no']}考场为{standard_seats}人，请确认是否为特殊安排。",
                })
                break

    return _build_seat_precheck_result(errors, warnings, stats, records, standard_seats)


def _build_seat_precheck_result(errors, warnings, stats, records, standard_seats: int) -> dict:
    room_count = len(stats)
    student_count = sum(item["count"] for item in stats)
    underfilled = [item for item in stats if item["count"] < standard_seats]
    return {
        "ok": len(errors) == 0,
        "standard_seats": standard_seats,
        "summary": {
            "room_count": room_count,
            "student_count": student_count,
            "full_room_count": room_count - len(underfilled),
            "underfilled_room_count": len(underfilled),
            "underfilled_rooms": underfilled,
            "room_range": f"1-{room_count}" if room_count else "",
        },
        "errors": errors,
        "warnings": warnings,
        "stats": stats,
        "records": records,
    }


def generate_seat_label_precheck_report(precheck: dict) -> BytesIO:
    output = BytesIO()
    summary = precheck.get("summary", {})
    summary_rows = [
        {"项目": "预检结果", "值": "通过" if precheck.get("ok") else "未通过"},
        {"项目": "考场数", "值": summary.get("room_count", 0)},
        {"项目": "考生数", "值": summary.get("student_count", 0)},
        {"项目": "满员考场数", "值": summary.get("full_room_count", 0)},
        {"项目": "不满员考场数", "值": summary.get("underfilled_room_count", 0)},
        {"项目": "每考场标准座位数", "值": precheck.get("standard_seats", 30)},
    ]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="摘要")
        pd.DataFrame(precheck.get("errors") or [{"message": "无错误"}]).to_excel(writer, index=False, sheet_name="错误")
        pd.DataFrame(precheck.get("warnings") or [{"message": "无警告"}]).to_excel(writer, index=False, sheet_name="警告")
        pd.DataFrame(precheck.get("stats") or []).to_excel(writer, index=False, sheet_name="考场统计")
        for ws in writer.book.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 48)
    output.seek(0)
    return output


def generate_seat_labels_pdf_v2(
    layout_mode: str,
    content_mode: str,
    border_mode: str = "print",
    exam_badge: str = "",
    num_rooms: int | None = None,
    num_seats: int = 30,
    cols: int = 3,
    rows: int = 10,
    font_size: int = 40,
    roster_precheck: dict | None = None,
    id_column: str | None = None,
) -> BytesIO:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise RuntimeError("缺少 reportlab 依赖，请先安装 requirements.txt") from exc

    if layout_mode not in {"stack_cut", "room_page"}:
        raise ValueError("生成模式无效")
    if content_mode not in {"numbers", "roster"}:
        raise ValueError("内容类型无效")
    if border_mode not in {"print", "none"}:
        raise ValueError("边框模式无效")
    exam_badge = (exam_badge or "").strip()
    if cols != 3 or rows != 10:
        raise ValueError("桌贴版式固定为 3 列 x 10 行")
    if num_seats < 1:
        raise ValueError("座位数必须大于 0")

    records = {}
    if content_mode == "roster":
        if not roster_precheck or not roster_precheck.get("ok"):
            raise ValueError("名单预检未通过")
        records = roster_precheck.get("records", {})
        num_rooms = roster_precheck.get("summary", {}).get("room_count", 0)
        num_seats = roster_precheck.get("standard_seats", num_seats)
        if not id_column:
            raise ValueError("请选择编号字段")
    if not num_rooms or num_rooms < 1:
        raise ValueError("考场数量必须大于 0")

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    page_width, page_height = A4
    cell_width = page_width / cols
    cell_height = page_height / rows
    label_inset = 2 * mm
    label_width = cell_width - label_inset * 2
    label_height = cell_height - label_inset * 2
    print_safe_margin = 6 * mm
    labels_per_page = cols * rows

    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)

    def label_xy(index: int):
        col = index // rows
        row = index % rows
        cell_x = col * cell_width
        cell_y = page_height - (row + 1) * cell_height
        x = cell_x + label_inset
        y = cell_y + label_inset
        return x, y, col, row

    def fit_font_size(text: str, width: float, font_name: str, max_size: int, min_size: int):
        size = max_size
        while size > min_size and pdf.stringWidth(text, font_name, size) > width:
            size -= 0.5
        return size

    def draw_fit_center(text: str, x: float, y: float, width: float, font_name: str, max_size: int, min_size: int):
        size = fit_font_size(text, width, font_name, max_size, min_size)
        pdf.setFont(font_name, size)
        pdf.drawCentredString(x, y, text)

    def draw_fit_left(text: str, x: float, y: float, width: float, font_name: str, max_size: int, min_size: int):
        size = fit_font_size(text, width, font_name, max_size, min_size)
        pdf.setFont(font_name, size)
        pdf.drawString(x, y, text)

    def draw_fit_right(text: str, right_x: float, y: float, width: float, font_name: str, max_size: int, min_size: int):
        size = fit_font_size(text, width, font_name, max_size, min_size)
        pdf.setFont(font_name, size)
        pdf.drawRightString(right_x, y, text)

    def draw_label(index: int, room_no: int, seat_no: int):
        x, y, col, row = label_xy(index)
        if border_mode == "print":
            pdf.setLineWidth(0.8)
            pdf.setStrokeColor(colors.black)
            pdf.rect(x, y, label_width, label_height)

        record = records.get((room_no, seat_no)) if content_mode == "roster" else True
        if not record:
            return

        title = f"{room_no}-{seat_no:02d}"
        content_left = x + 2 * mm
        content_right = x + label_width - 2 * mm
        if border_mode == "none":
            if col == 0:
                content_left = max(content_left, print_safe_margin)
            if col == cols - 1:
                content_right = min(content_right, page_width - print_safe_margin)
        content_width = content_right - content_left

        if content_mode == "numbers":
            draw_fit_center(title, content_left + content_width / 2, y + label_height / 2 - font_size / 2.8, content_width, "Helvetica-Bold", font_size, 12)
            return

        text_x = content_left
        info_width = content_width
        content_bottom = y
        content_top = y + label_height
        if border_mode == "none":
            if row == 0:
                content_top = min(content_top, page_height - print_safe_margin)
            if row == rows - 1:
                content_bottom = max(content_bottom, print_safe_margin)

        section_gap = 2 * mm
        right_width = info_width * 0.32
        left_width = info_width - right_width - section_gap
        right_left = content_right - right_width
        right_center_x = right_left + right_width / 2
        content_center_y = content_bottom + (content_top - content_bottom) / 2

        left_lines = []
        if exam_badge:
            left_lines.append((exam_badge, "STSong-Light", 10, 8))
        left_lines.extend([
            (f"第{room_no}考场", "STSong-Light", 10, 8),
            (f"姓名：{record['name']}", "STSong-Light", 10, 6),
            (f"{id_column}：{record['identifier']}", "STSong-Light", 9, 5.5),
        ])
        line_gap = 2.2
        line_sizes = [fit_font_size(text, left_width, font, max_size, min_size) for text, font, max_size, min_size in left_lines]
        line_heights = [size * 1.1 for size in line_sizes]
        total_left_height = sum(line_heights) + line_gap * (len(line_heights) - 1)
        left_top = content_center_y + total_left_height / 2
        current_y = left_top
        for (text, font, _, _), size, height in zip(left_lines, line_sizes, line_heights):
            baseline = current_y - height * 0.82
            pdf.setFont(font, size)
            pdf.drawString(text_x, baseline, text)
            current_y -= height + line_gap

        seat_text = f"{seat_no:02d}"
        seat_size = fit_font_size(seat_text, right_width, "Helvetica-Bold", 60, 36)
        seat_y = content_center_y - seat_size * 0.34
        pdf.setFont("Helvetica-Bold", seat_size)
        pdf.drawCentredString(right_center_x, seat_y, seat_text)

    if layout_mode == "stack_cut":
        room_batches = math.ceil(num_rooms / labels_per_page)
        for batch in range(room_batches):
            start_room = batch * labels_per_page + 1
            rooms_this_batch = min(num_rooms - batch * labels_per_page, labels_per_page)
            for seat_no in range(1, num_seats + 1):
                for idx in range(rooms_this_batch):
                    draw_label(idx, start_room + idx, seat_no)
                pdf.showPage()
    else:
        seat_pages = math.ceil(num_seats / labels_per_page)
        for room_no in range(1, num_rooms + 1):
            for page in range(seat_pages):
                start_seat = page * labels_per_page + 1
                seats_this_page = min(num_seats - page * labels_per_page, labels_per_page)
                for idx in range(seats_this_page):
                    draw_label(idx, room_no, start_seat + idx)
                pdf.showPage()

    pdf.save()
    output.seek(0)
    return output


def merge_excel_sheets(file_bytes: bytes) -> BytesIO:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    merged_sheet_name = "合并数据"
    if merged_sheet_name in wb.sheetnames:
        del wb[merged_sheet_name]

    merged_sheet = wb.create_sheet(title=merged_sheet_name)
    header_data = []
    row_num = 1

    for sheet_name in list(wb.sheetnames):
        if sheet_name == merged_sheet_name:
            continue
        sheet = wb[sheet_name]
        if sheet.max_row < 1:
            continue

        for row in range(1, sheet.max_row + 1):
            current_row_data = [
                sheet.cell(row=row, column=col).value if sheet.cell(row=row, column=col).value is not None else ""
                for col in range(1, sheet.max_column + 1)
            ]
            if row == 1 and not header_data:
                header_data = current_row_data
                for col, value in enumerate(current_row_data, start=1):
                    merged_sheet.cell(row=row_num, column=col, value=value)
                row_num += 1
                continue
            if current_row_data == header_data:
                continue
            if all(str(cell).strip() == "" for cell in current_row_data):
                continue
            for col, value in enumerate(current_row_data, start=1):
                merged_sheet.cell(row=row_num, column=col, value=value)
            row_num += 1

    sheet_names = wb.sheetnames
    sheet_names.remove(merged_sheet_name)
    sheet_names.insert(0, merged_sheet_name)
    wb._sheets = [wb[name] for name in sheet_names]

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output


def extract_grade_from_filename(filename: str) -> tuple[str, int]:
    match = re.search(r"(\d{4})届", filename)
    if match:
        grade_year = int(match.group(1))
        return f"{grade_year}届", grade_year
    return os.path.splitext(filename)[0], 9999


def read_dbf_files(file_paths: list[str]) -> pd.DataFrame:
    try:
        from dbfread import DBF
    except ImportError as exc:
        raise RuntimeError("缺少 dbfread 依赖，请先安装 requirements.txt") from exc

    all_dbf_data = []
    for file_path in file_paths:
        dbf_table = DBF(file_path, encoding="gbk")
        all_dbf_data.append(pd.DataFrame(iter(dbf_table)))
    if not all_dbf_data:
        raise ValueError("未找到 DBF 成绩文件")
    return pd.concat(all_dbf_data, ignore_index=True).drop_duplicates()


def clean_score_data(score_df: pd.DataFrame, graduate_id_list: list) -> set:
    _require_columns(score_df, {SCORE_ID_COL, SCORE_VALUE_COL}, "成绩表")
    cleaned = score_df.copy()
    cleaned[SCORE_VALUE_COL] = pd.to_numeric(cleaned[SCORE_VALUE_COL], errors="coerce")
    valid_score_df = cleaned[
        (cleaned[SCORE_VALUE_COL].notna()) & (cleaned[SCORE_VALUE_COL] >= PASS_SCORE)
    ].copy()
    valid_score_df = valid_score_df[valid_score_df[SCORE_ID_COL].isin(graduate_id_list)]
    return set(valid_score_df[SCORE_ID_COL].drop_duplicates().tolist())


def calculate_single_level_detail(graduate_df: pd.DataFrame, pass_ids: set, exam_type: str, level: str, grade: str) -> pd.DataFrame:
    total_all = len(graduate_df)
    pass_all = len(pass_ids)
    rate_all = round((pass_all / total_all) * 100, 2) if total_all > 0 else 0.0

    valid_graduate_df = graduate_df[~graduate_df[GRADUATE_MAJOR_COL].isin(EXCLUDE_MAJORS)]
    total_valid = len(valid_graduate_df)
    valid_ids = set(valid_graduate_df[GRADUATE_ID_COL].tolist())
    pass_valid = len(pass_ids & valid_ids)
    rate_valid = round((pass_valid / total_valid) * 100, 2) if total_valid > 0 else 0.0

    major_stats = []
    for major, major_df in graduate_df.groupby(GRADUATE_MAJOR_COL):
        major_ids = set(major_df[GRADUATE_ID_COL].tolist())
        major_pass = len(pass_ids & major_ids)
        major_total = len(major_df)
        major_rate = round((major_pass / major_total) * 100, 2) if major_total > 0 else 0.0
        major_stats.append({
            "届别": grade,
            "培养层次": level,
            "专业": major,
            "专业总人数": major_total,
            f"{exam_type}通过人数": major_pass,
            f"{exam_type}通过率(%)": major_rate,
        })

    overall_all = {
        "届别": grade,
        "培养层次": level,
        "专业": "全部该层次（含所有专业）",
        "专业总人数": total_all,
        f"{exam_type}通过人数": pass_all,
        f"{exam_type}通过率(%)": rate_all,
    }
    overall_valid = {
        "届别": grade,
        "培养层次": level,
        "专业": "有效整体（排除指定专业）",
        "专业总人数": total_valid,
        f"{exam_type}通过人数": pass_valid,
        f"{exam_type}通过率(%)": rate_valid,
    }
    return pd.DataFrame([overall_all, overall_valid] + major_stats)


def _write_single_grade_level_excel(archive: zipfile.ZipFile, grade: str, level: str, cet4_result: pd.DataFrame, cet6_result: pd.DataFrame):
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_title, df in (("四级通过率统计", cet4_result), ("六级通过率统计", cet6_result)):
        ws = wb.create_sheet(title=sheet_title)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = 15
    wb.save(output)
    archive.writestr(f"{grade}{level}四六级通过率统计.xlsx", output.getvalue())


def calculate_cet_pass_rates(graduate_files: list[tuple[str, bytes]], cet4_paths: list[str], cet6_paths: list[str]) -> BytesIO:
    if not graduate_files:
        raise ValueError("请上传至少一个毕业生届别 Excel")
    if not cet4_paths or not cet6_paths:
        raise ValueError("请分别上传四级和六级 DBF 成绩文件")

    cet4_score_df = read_dbf_files(cet4_paths)
    cet6_score_df = read_dbf_files(cet6_paths)
    all_major_trend_data = []
    all_grade_level_data = []

    zip_output = BytesIO()
    with zipfile.ZipFile(zip_output, "w", zipfile.ZIP_DEFLATED) as archive:
        for filename, file_bytes in graduate_files:
            graduate_df = pd.read_excel(BytesIO(file_bytes))
            _require_columns(graduate_df, {GRADUATE_ID_COL, GRADUATE_MAJOR_COL, GRADUATE_LEVEL_COL}, f"{filename}")
            graduate_df = graduate_df.drop_duplicates(subset=[GRADUATE_ID_COL]).reset_index(drop=True)
            grade_name, grade_year = extract_grade_from_filename(filename)

            for level in TARGET_LEVELS:
                level_df = graduate_df[graduate_df[GRADUATE_LEVEL_COL] == level].copy()
                level_total = len(level_df)
                if level_total == 0:
                    all_grade_level_data.append({
                        "届别": grade_name,
                        "培养层次": level,
                        "总人数": 0,
                        "四级通过人数": 0,
                        "四级通过率(%)": 0.0,
                        "六级通过人数": 0,
                        "六级通过率(%)": 0.0,
                        "有效总人数": 0,
                        "有效四级通过率(%)": 0.0,
                        "有效六级通过率(%)": 0.0,
                    })
                    continue

                level_id_list = level_df[GRADUATE_ID_COL].tolist()
                cet4_pass_ids = clean_score_data(cet4_score_df, level_id_list)
                cet6_pass_ids = clean_score_data(cet6_score_df, level_id_list)
                cet4_detail_df = calculate_single_level_detail(level_df, cet4_pass_ids, "四级", level, grade_name)
                cet6_detail_df = calculate_single_level_detail(level_df, cet6_pass_ids, "六级", level, grade_name)
                _write_single_grade_level_excel(archive, grade_name, level, cet4_detail_df, cet6_detail_df)

                valid_df = level_df[~level_df[GRADUATE_MAJOR_COL].isin(EXCLUDE_MAJORS)]
                valid_ids = set(valid_df[GRADUATE_ID_COL].tolist())
                cet4_pass_all = len(cet4_pass_ids)
                cet6_pass_all = len(cet6_pass_ids)
                all_grade_level_data.append({
                    "届别": grade_name,
                    "培养层次": level,
                    "总人数": level_total,
                    "四级通过人数": cet4_pass_all,
                    "四级通过率(%)": round((cet4_pass_all / level_total) * 100, 2),
                    "六级通过人数": cet6_pass_all,
                    "六级通过率(%)": round((cet6_pass_all / level_total) * 100, 2),
                    "有效总人数": len(valid_df),
                    "有效四级通过率(%)": round((len(cet4_pass_ids & valid_ids) / len(valid_df)) * 100, 2) if len(valid_df) else 0.0,
                    "有效六级通过率(%)": round((len(cet6_pass_ids & valid_ids) / len(valid_df)) * 100, 2) if len(valid_df) else 0.0,
                })

                for major, major_df in level_df.groupby(GRADUATE_MAJOR_COL):
                    major_ids = set(major_df[GRADUATE_ID_COL].tolist())
                    major_total = len(major_df)
                    all_major_trend_data.append({
                        "培养层次": level,
                        "专业": major,
                        "届别": grade_name,
                        "届别年份": grade_year,
                        "总人数": major_total,
                        "四级通过人数": len(cet4_pass_ids & major_ids),
                        "四级通过率(%)": round((len(cet4_pass_ids & major_ids) / major_total) * 100, 2) if major_total else 0.0,
                        "六级通过人数": len(cet6_pass_ids & major_ids),
                        "六级通过率(%)": round((len(cet6_pass_ids & major_ids) / major_total) * 100, 2) if major_total else 0.0,
                    })

        trend_raw_df = pd.DataFrame(all_major_trend_data)
        if not trend_raw_df.empty:
            all_grades_sorted = sorted(set(trend_raw_df["届别"]), key=lambda x: extract_grade_from_filename(x)[1])
            trend_data = []
            for (level, major), group_df in trend_raw_df.groupby(["培养层次", "专业"]):
                row = {"培养层次": level, "专业": major}
                for grade in all_grades_sorted:
                    grade_df = group_df[group_df["届别"] == grade]
                    if len(grade_df) > 0:
                        row[f"{grade}-总人数"] = grade_df["总人数"].iloc[0]
                        row[f"{grade}-四级通过率(%)"] = grade_df["四级通过率(%)"].iloc[0]
                        row[f"{grade}-六级通过率(%)"] = grade_df["六级通过率(%)"].iloc[0]
                    else:
                        row[f"{grade}-总人数"] = "-"
                        row[f"{grade}-四级通过率(%)"] = "-"
                        row[f"{grade}-六级通过率(%)"] = "-"
                trend_data.append(row)
            trend_df = pd.DataFrame(trend_data).sort_values(by=["培养层次", "专业"])
        else:
            trend_df = pd.DataFrame()

        summary_df = pd.DataFrame(all_grade_level_data).sort_values(by=["届别", "培养层次"])
        summary_output = BytesIO()
        with pd.ExcelWriter(summary_output, engine="openpyxl") as writer:
            trend_df.to_excel(writer, sheet_name="专业通过率趋势表", index=False)
            summary_df.to_excel(writer, sheet_name="整体通过率汇总表", index=False)
            for ws in writer.book.worksheets:
                for cell in ws[1]:
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal="center")
                for column in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                    ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 24)
        archive.writestr("多届专业通过率趋势分析表.xlsx", summary_output.getvalue())

    zip_output.seek(0)
    return zip_output
